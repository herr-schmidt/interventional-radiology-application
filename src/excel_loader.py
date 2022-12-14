from enum import Enum
from openpyxl import load_workbook

from src.model import Patient


class MainSheetNotFound(Exception):
    def __init__(self, message):
        super().__init__(message)


class InvalidRow(Exception):
    def __init__(self, message):
        super().__init__(message)


class ExcelIndex(Enum):
    NAME = 0
    SURNAME = 1
    SERVICES = 2
    ANESTHESIA = 3
    INFECTIOUS = 4
    LIST_INSERTION_DATE = 5


class ExcelLoader:
    def __init__(self):
        pass

    @staticmethod
    def has_none_fields(row) -> bool:
        return (row[ExcelIndex.NAME.value] is None
                or row[ExcelIndex.SURNAME.value] is None
                or row[ExcelIndex.SERVICES.value] is None
                or row[ExcelIndex.ANESTHESIA.value] is None
                or row[ExcelIndex.INFECTIOUS.value] is None
                or row[ExcelIndex.LIST_INSERTION_DATE.value] is None)

    def find_main_sheet(self, workbook):
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            workbook.active = workbook[sheet_name]
            # we check for the right header
            if list(workbook.active.values)[0][0] == "Nome":
                return sheet_name
        raise MainSheetNotFound("Main sheet not found")

    def load_patients(self, xlsx_file_name) -> list[Patient]:
        wb = load_workbook(xlsx_file_name)
        wb.iso_dates = True
        main_sheet = self.find_main_sheet(wb)
        active_sheet = wb[main_sheet]
        patients = []
        rows = list(active_sheet.values)
        for row in rows[1:]:  # skip header
            if ExcelLoader.has_none_fields(row):
                raise InvalidRow("Row contains at least one None value among its fields.")

            name = row[ExcelIndex.NAME.value].strip()
            surname = row[ExcelIndex.SURNAME.value].strip()
            services = row[ExcelIndex.SERVICES.value].strip()
            anesthesia = row[ExcelIndex.ANESTHESIA.value].lower() == "true"
            infectious = row[ExcelIndex.INFECTIOUS.value].lower() == "true"
            list_insertion_date = row[ExcelIndex.LIST_INSERTION_DATE.value]

            patients.append(Patient(name,
                                    surname,
                                    services,
                                    anesthesia,
                                    infectious,
                                    list_insertion_date))

        return patients
